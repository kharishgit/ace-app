from rest_framework import (viewsets, filters, status)
from django.shortcuts import render
from rest_framework.decorators import api_view
from accounts.models import Course,Subject,Module,Topic,SubTopic,Batch,Branch,User,FacultyCourseAddition
from rest_framework.response import Response
from .serializers import CourseSearch
from rest_framework.filters import SearchFilter,OrderingFilter
from rest_framework import generics
from rest_framework.pagination import PageNumberPagination
from course.serializers import SubjectSerializer,ModuleSerializer,TopicSerializer,SubTopicSerializer,BatchSerializer,BranchSerializer
from accounts.api.serializers import UserSerializer,facultyviewDetails,FacultyCourseAdditionsss
from .resources import export_to_pdf
from .utils import generate_pdf
# export_to_pdfs #for creating pdf common function
#excel fomate of each tables 9/2/2023
from django.http import HttpResponse, HttpResponseBadRequest
from .resources import FacultyCourseResource, CourseResource, BatchResource,SubjectResource,ModuleResource,TopicResource,SubTopicResource,BranchResource,SuperAdminResource,FacultyResource
from course.models import Holidays,SpecialHoliday,Level,Batch,Branch
from django.http import HttpResponse, JsonResponse
from django.core.exceptions import ValidationError
from django.views.generic import View

#pagination setup (query_param like search/coursecourseslist/page_size={num}
class MyPagination(PageNumberPagination):
    page_size = 6
    page_size_query_param = 'page_size'  #query parameter like page_size=5(number of data)
    max_page_size = 100

# Create your views here.
#course pagination,ordering and search and pagination
class CoursePagination(generics.ListAPIView):
    queryset = Course.objects.all()
    serializer_class = CourseSearch
    filter_backends = [SearchFilter,OrderingFilter]
    ordering_fields = ['name',]    #ordering with name like name ascending and descending
    search_fields = ['^name', 'name','active']   #start with name,name letters
    pagination_class = MyPagination


class BatchPagination(generics.ListAPIView):
    queryset = Batch.objects.all()
    serializer_class = BatchSerializer
    filter_backends = [SearchFilter,OrderingFilter]
    ordering_fields = ['name','course__name','branch__name','start_date','end_date']
    search_fields = ['^name','name','start_date','end_date','active','strength']
    pagination_class = MyPagination


#subjecy pagination,ordering and search and pagination
class SubjectPagination(generics.ListAPIView):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    filter_backends = [SearchFilter,OrderingFilter]
    ordering_fields = ['name']      #ordering with name like name ascending and descending
    search_fields = ['^name','name','active']    #start with name,name letters
    pagination_class = MyPagination

#Module pagination,ordering and search and pagination
class ModulePagination(generics.ListAPIView):
    queryset = Module.objects.all()
    serializer_class = ModuleSerializer
    filter_backends = [SearchFilter,OrderingFilter]
    ordering_fields = ['name']
    search_fields = ['^name','name','active']
    pagination_class = MyPagination

class TopicPagination(generics.ListAPIView):
    queryset = Topic.objects.all()
    serializer_class = TopicSerializer
    filter_backends = [SearchFilter,OrderingFilter]
    ordering_fields = ['name']
    search_fields = ['^name','name','active']
    pagination_class = MyPagination

class SubtopicPagination(generics.ListAPIView):
    queryset = SubTopic.objects.all()
    serializer_class = SubTopicSerializer
    filter_backends = [SearchFilter,OrderingFilter]
    ordering_fields = ['name']
    search_fields = ['^name','name','active']
    pagination_class = MyPagination



class BranchPagination(generics.ListAPIView):
    queryset = Branch.objects.all()
    serializer_class = BranchSerializer
    filter_backends = [SearchFilter,OrderingFilter]
    ordering_fields = ['name','location','courses__name']
    search_fields = ['^name','name','location','active']
    pagination_class = MyPagination


# class BranchadminPagination(generics.ListAPIView):
#     queryset = BranchAdmin.objects.all()
#     serializer_class = BranchAdminSerializer
#     filter_backends = [SearchFilter,OrderingFilter]
#     ordering_fields = ['email','superadmin__username','branch__name']
#     search_fields = ['^email','email','superadmin','branch']
#     pagination_class = MyPagination


class SuperadminPagination(generics.ListAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    filter_backends = [SearchFilter,OrderingFilter]
    ordering_fields = ['name','mobile']
    search_fields = ['^username','username','email']
    pagination_class = MyPagination


#faculty search on name,subjct,course,module,topic,
from django.db.models import Q
from django.shortcuts import get_object_or_404
# from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, generics

from accounts.models import Faculty
from accounts.api.serializers import FacultySerializerforisverifiedandNOT
from django.db.models import Q
from rest_framework.pagination import PageNumberPagination
class FacultyListView(generics.ListAPIView):
    serializer_class = facultyviewDetails
    filter_backends = [filters.SearchFilter]
    search_fields = ['user__username', 'user__email', 'user__mobile']
    pagination_class=MyPagination

    def get_queryset(self):
        queryset = Faculty.objects.all().order_by('id')
        print(queryset.count())

        # Apply search filter
        search_query = self.request.query_params.get('search', None)
        if search_query:
            queryset = queryset.filter(Q(user__username__icontains=search_query) |
                                       Q(user__email__icontains=search_query) |
                                       Q(user__mobile__icontains=search_query))

        # Apply filter based on FacultyCourseAddition's fields
        course = self.request.query_params.get('course', None)
        category = self.request.query_params.get('category', None)
        level = self.request.query_params.get('level', None)
        subject = self.request.query_params.get('subject', None)
        module = self.request.query_params.get('module', None)
        topic = self.request.query_params.get('topic', None)

        if course or category or level or subject or module or topic:
            queryset = queryset.filter(facultycourseaddition__course__icontains=course,
                                       facultycourseaddition__category__icontains=category,
                                       facultycourseaddition__level__icontains=level,
                                       facultycourseaddition__subject__name__icontains=subject,
                                       facultycourseaddition__module__name__icontains=module,
                                       facultycourseaddition__topic__name__icontains=topic)
        queryset = queryset.order_by('id')
        return queryset





# class FacultyListView(generics.ListAPIView):
#     serializer_class = facultyviewDetails
#     filter_backends = [filters.SearchFilter]
#     search_fields = ['user__username', 'user__email', 'user__mobile',
#                      'facultycourseaddition__category__name',
#                      'facultycourseaddition__level__name',
#                      'facultycourseaddition__course__name',
#                      'facultycourseaddition__subject__name',
#                      'facultycourseaddition__module__name',
#                      'facultycourseaddition__topic__name']

#     def get_queryset(self):
#         queryset = Faculty.objects.all()

#         # Apply search filter
#         search_query = self.request.query_params.get('search', None)
#         if search_query:
#             queryset = queryset.filter(Q(user__username__icontains=search_query) |
#                                        Q(user__email__icontains=search_query) |
#                                        Q(user__mobile__icontains=search_query) |
#                                        Q(facultycourseaddition__category__name__icontains=search_query) |
#                                        Q(facultycourseaddition__level__name__icontains=search_query) |
#                                        Q(facultycourseaddition__course__name__icontains=search_query) |
#                                        Q(facultycourseaddition__subject__name__icontains=search_query) |
#                                        Q(facultycourseaddition__module__name__icontains=search_query) |
#                                        Q(facultycourseaddition__topic__name__icontains=search_query))

#         return queryset

# class FacultyListView(generics.ListAPIView):
#     serializer_class = facultyviewDetails
#     filter_backends = [filters.SearchFilter]
#     # filterset_fields = ['whatsapp_contact_number','expected_salary','course', 'subject__name', 'module__name', 'topic__name']
#     search_fields = ['user__username', 'user__email', 'user__mobile']

#     def get_queryset(self):
#         queryset = Faculty.objects.all()

#         # Apply search filter
#         search_query = self.request.query_params.get('search', None)
#         if search_query:
#             queryset = queryset.filter(Q(user__username__icontains=search_query) |
#                                        Q(user__email__icontains=search_query) |
#                                        Q(user__mobile__icontains=search_query))



#         # queryset = self.filter_queryset(queryset)
#         # print(str(queryset.query))
#         return queryset

# class FacultyListViewss(generics.ListAPIView):
#     serializer_class = FacultyCourseAdditionsss
#     filter_backends = [filters.SearchFilter]
#     filterset_fields = ['category__name ','level__name','course__name' ,'subject__name', 'module__name', 'topic__name']
#     search_fields = ['category__name', 'level__name', 'course__name','subject__name','module__name','topic__name']

#     def get_queryset(self):
#         queryset = FacultyCourseAddition.objects.all()

#         # Apply search filter
#         search_query = self.request.query_params.get('search', None)
#         if search_query:
#             queryset = queryset.filter(Q(category__name__icontains=search_query) |
#                                        Q(level__name__icontains=search_query) |
#                                     #    Q(user__mobile__icontains=search_query)|
#                                        Q(course__name__icontains=search_query)|
#                                        Q(subject__name__icontains=search_query)|
#                                        Q(module__name__icontains=search_query)|
#                                        Q(topic__name__icontains=search_query))


#         # queryset = self.filter_queryset(queryset)
#         # print(str(queryset.query))
#         return queryset



#excel fomate of each tables 9/2/2023
# views.py

# def export_COURSE_to_excel(request):
#     course_resource = CourseResource()
#     dataset = course_resource.export()
#     response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
#     response['Content-Disposition'] = 'attachment; filename="courses.xls"'
#     return response

def export_BATCH_to_excel(request):
    batch_resource = BatchResource()
    dataset = batch_resource.export()
    response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="batches.xls"'
    return response


def export_SUBJECT_to_excel(request):
    batch_resource = SubjectResource()
    dataset = batch_resource.export()
    response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="batches.xls"'
    return response

def export_MODULE_to_excel(request):
    batch_resource = ModuleResource()
    dataset = batch_resource.export()
    response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="batches.xls"'
    return response



def export_TOPIC_to_excel(request):
    batch_resource = TopicResource()
    dataset = batch_resource.export()
    response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="batches.xls"'
    return response

def export_SUBTOPIC_to_excel(request):
    batch_resource = SubTopicResource()
    dataset = batch_resource.export()
    response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="batches.xls"'
    return response

def export_BRANCH_to_excel(request):
    batch_resource = BranchResource()
    dataset = batch_resource.export()
    response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="batches.xls"'
    return response

# def export_BRANCHADMIN_to_excel(request):
#     batch_resource = BrachAdminResource()
#     dataset = batch_resource.export()
#     response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
#     response['Content-Disposition'] = 'attachment; filename="batches.xls"'
#     return response

def export_SUPERADMIN_to_excel(request):
    batch_resource = SuperAdminResource()
    dataset = batch_resource.export()
    response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="batches.xls"'
    return response

#commmon function
def export_to_excel(request, resource_class):
    resource_class_str = resource_class + 'Resource'
    print(resource_class_str)
    print(resource_class)
    resource = eval(resource_class_str)()
    print(resource)
    dataset = resource.export()
    response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="data.xls"'
    return response

def export_FACAULTY_to_excel(request):
    batch_resource = FacultyResource()
    dataset = batch_resource.export()
    response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="batches.xls"'
    return response


##################### end data to excel ###########################



################# start pdf function ##################
def export_courses_pdf(request):
    courses = Course.objects.all()
    fields = ['id','name']
    file_name = 'courses.pdf'
    return export_to_pdf(courses, fields, file_name)

def export_batch_pdf(request):
    courses = Batch.objects.all()
    fields = ['id','name','start_date','end_date','course','branch']
    file_name = 'courses.pdf'
    return export_to_pdf(courses, fields, file_name)

def export_subject_pdf(request):
    courses = Subject.objects.all()
    fields = ['name','course']
    file_name = 'courses.pdf'
    return export_to_pdf(courses, fields, file_name)

def export_MODULE_pdf(request):
    courses = Module.objects.all()
    fields = ['id','name','subject']
    file_name = 'courses.pdf'
    return export_to_pdf(courses, fields, file_name)

def export_TOPIC_pdf(request):
    courses = Topic.objects.all()
    fields = ['id','name','module','time_needed']
    file_name = 'courses.pdf'
    return export_to_pdf(courses, fields, file_name)

def export_SUBTOPIC_pdf(request):
    courses = SubTopic.objects.all()
    fields = ['id','name','topic']
    file_name = 'courses.pdf'
    return export_to_pdf(courses, fields, file_name)

# def export_BRANCH_pdf(request):
#     courses = Branch.objects.all()
#     fields = ['id','name','location','courses']
#     file_name = 'courses.pdf'
#     return export_to_pdf(courses, fields, file_name)
def export_BRANCH_pdf(request):
    branches = Branch.objects.all()
    file_name = "branch_list.pdf"
    response = export_to_pdfs(branches, file_name)
    return response

# def export_BRANCHADMIN_pdf(request):
#     branches = BranchAdmin.objects.all()
#     fields = ['id','email','superadmin','branch']
#     file_name = 'BranchAdmin.pdf'
#     return export_to_pdf(branches, fields, file_name)


def export_SUPERADMIN_pdf(request):
    branches = User.objects.all()
    fields = ['id','username','email','mobile','joined_date']
    file_name = 'SuperAdmin.pdf'
    return export_to_pdf(branches, fields, file_name)




##################common pdf#################3
# def common_pdf_formate(request, model_name):
#     if model_name == 'course':
#         queryset = Course.objects.all()
#         fields = ['id', 'name','description']
#         file_name = 'courses.pdf'
#     elif model_name == 'batch':
#         queryset = Batch.objects.all()
#         fields = ['id', 'name', 'start_date', 'end_date', 'course', 'branch']
#         file_name = 'batches.pdf'
#     elif model_name == 'subject':
#         queryset = Subject.objects.all()
#         fields = ['id', 'name', 'course']
#         file_name = 'subjects.pdf'
#     elif model_name == 'module':
#         queryset = Module.objects.all()
#         fields = ['id', 'name', 'subject']
#         file_name = 'modules.pdf'
#     elif model_name == 'topic':
#         queryset = Topic.objects.all()
#         fields = ['id', 'name', 'module', 'time_needed']
#         file_name = 'topics.pdf'
#     elif model_name == 'subtopic':
#         queryset = SubTopic.objects.all()
#         fields = ['id', 'name', 'topic']
#         file_name = 'subtopics.pdf'
#     else:
#         return HttpResponseBadRequest("Invalid model name")

#     return export_to_pdf(queryset, fields, file_name)




def common_pdf_formate(request, model_name):
    if model_name == 'course':
        queryset = Course.objects.all()
        fields = ['id', 'name','description','batch_type','active','year']
        file_name = 'courses.pdf'
        header_text = 'Course List'
    elif model_name == 'batch':
        queryset = Batch.objects.all()
        fields = ['id', 'name', 'start_date', 'end_date', 'course', 'branch']
        file_name = 'batches.pdf'
        header_text = 'Batches'
    elif model_name == 'subject':
        queryset = Subject.objects.all()
        fields = ['id', 'name', 'course','description','priority','active']
        file_name = 'subjects.pdf'
        header_text = 'Subjects'
    elif model_name == 'module':
        queryset = Module.objects.all()
        fields = ['id', 'name', 'subject','description','priority','active']
        file_name = 'modules.pdf'
        header_text = 'Modules'
    elif model_name == 'topic':
        queryset = Topic.objects.all()
        fields = ['id', 'name', 'module','description','time_needed','active']
        file_name = 'topics.pdf'
        header_text = 'Topics'
    elif model_name == 'subtopic':
        queryset = SubTopic.objects.all()
        fields = ['id', 'name', 'topic','description','time_needed','active']
        file_name = 'subtopics.pdf'
        header_text = 'Subtopics'
    elif model_name == 'branch':
        queryset = Branch.objects.all()
        fields = ['id', 'name', 'location','active']
        file_name = 'Bracnh.pdf'
        header_text = 'Branch'

    else:
        return HttpResponseBadRequest("Invalid model name")

    response = export_to_pdf(queryset, fields, file_name, header_text)
    return response


# from django.db import models

# def get_table_data(model):
#     print("IIIIIIIIIIII")
#     # Get all fields in the model
#     fields = model._meta.get_fields()
#     print("IIIIIIIIIIIIDDD")
#     # Extract the names of fields that are not related fields or ManyToMany fields
#     headers = [field.name for field in fields
#                if isinstance(field, models.Field) and not (field.one_to_many or field.one_to_one or field.many_to_many)]

#     # Get all instances of the model
#     instances = model.objects.all()

#     # Extract the data for each instance
#     data = [[getattr(instance, header) for header in headers] for instance in instances]

#     return headers, data

# from django.http import HttpResponse
# from django.template.loader import get_template
# from course.models import Course
# from xhtml2pdf import pisa
# import io
# from django.utils import timezone
# from django.apps import apps

# def table_to_pdf(request, model_name):
#     # Get the model dynamically based on the model_name parameter
#     model = apps.get_model(app_label='course', model_name=model_name)

#     # Get the table headers and data for the model
#     headers, data = get_table_data(model)
#     current_datetime = timezone.now()
#     # Pass the headers and data to the template
#     context = {
#         'headers': headers,
#         'data': data,
#         'model':model_name,
#         'current_datetime': current_datetime,
#     }

#     # Render the template to html string
#     template = get_template('commonpdf.html')
#     html_string = template.render(context)

#     # Create a pdf file using the html string
#     result = io.BytesIO()
#     pdf = pisa.CreatePDF(io.StringIO(html_string), dest=result, encoding='utf-8')
#     if not pdf.err:
#         # Return the pdf file as a response
#         response = HttpResponse(result.getvalue(), content_type='application/pdf')
#         response['Content-Disposition'] = 'attachment; filename="{}.pdf"'.format(model_name)
#         return response
#     return HttpResponse('Error creating pdf file', status=400)


from django.db import models
from django.http import HttpResponse
from django.template.loader import get_template
from django.apps import apps
from xhtml2pdf import pisa
import io
from django.utils import timezone

# def get_table_data(model):
#     # Get all fields in the model
#     fields = model._meta.get_fields()

#     # Extract the names of fields that are not related fields or ManyToMany fields
#     headers = [field.name for field in fields
#                if isinstance(field, models.Field) and not (field.one_to_many or field.one_to_one or field.many_to_many)and field.name != 'photo' and field.name != 'description']

#     # Get all instances of the model
#     instances = model.objects.all()

#     # Extract the data for each instance
#     data = [[getattr(instance, header) for header in headers] for instance in instances]

#     return headers, data

def table_to_pdf(request, model_name):
    # Get the model dynamically based on the model_name parameter
    model = apps.get_model(app_label='course', model_name=model_name)

    # Get the table headers and data for the model
    headers, data = get_table_data(model)
    current_datetime = timezone.now()

    # Pass the headers and data to the template
    context = {
        'headers': headers,
        'data': data,
        'model':model_name,
        'current_datetime': current_datetime,
    }

    # Render the template to html string
    template = get_template('commonpdf.html')
    html_string = template.render(context)

    # Create a pdf file using the html string
    result = io.BytesIO()
    pdf = pisa.CreatePDF(io.StringIO(html_string), dest=result, encoding='utf-8')
    if not pdf.err:
        # Return the pdf file as a response
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="{}.pdf"'.format(model_name)
        return response
    return HttpResponse('Error creating pdf file', status=400)























#export data to excel
# from django.http import HttpResponse
# import openpyxl


# from django.http import HttpResponse
# import openpyxl

# def export_data_to_excel(request):
#     # Retrieve data from the database
#     data = Course.objects.all()

#     # Get the list of field names from the model
#     field_names = [field.name for field in Course._meta.get_fields()]

#     # Create the HttpResponse object with the appropriate content type
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=your_file_name.xlsx'

#     # Create the workbook and add a worksheet
#     workbook = openpyxl.Workbook()
#     worksheet = workbook.active

#     # Write the header row
#     worksheet.append(field_names)

#     # Write the data
#     for item in data:
#         row = [getattr(item, field_name) for field_name in field_names]
#         worksheet.append(row)

#     # Save the workbook to the response object
#     workbook.save(response)

#     return response




# def export_data_to_excel(request):
#     # Retrieve data from the database
#     # fields = Course._meta.get_fields()
#     fields = [field.name for field in Subject._meta.fields]


#     print(fields,'ddd')
#     data = Subject.objects.all()
#     print(data)
#     for c in data:
#         # print(c.course)
#         # print(c.name)
#         print(c,"CCCCCCCCCCCCCCC")

#     # Create the HttpResponse object with the appropriate content type
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=your_file_name.xlsx'

#     # Create the workbook and add a worksheet
#     workbook = openpyxl.Workbook()
#     worksheet = workbook.active

#     # Write the header row
#     # worksheet.append(["name",])
#     fields = [field.name for field in Subject._meta.fields]

#     b='name'
#     a='item.'
#     c=a+b
#     print(c,'ccc')
#     for x in fields:
#         b=x
#         a='item.'
#         c=a+b
#         worksheet.append([x])
#         for item in data:
#             # print(item.x)
#             worksheet.append([c,item.course.name])
#             print(worksheet,'kkk00')
#             print(workbook,'kkkkkkkk')
#             print(item,'items')
#         # print(x,"xx")
#         # print(item.x,'dddddddddddddddddd')
#     # Write the data
#     # for item in data:
#     #     worksheet.append([item.name,])

#     # Save the workbook to the response object
#     workbook.save(response)

#     return response


#shamil pdf to
# from django.http import FileResponse
# from django.template.loader import get_template
# from io import BytesIO
# from reportlab.pdfgen import canvas

# def export_data_to_pdf(request):
#     # Retrieve data from the database
#     data = Course.objects.all()

#     # Create the HttpResponse object with the appropriate content type
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename=your_file_name.pdf'

#     # Create the PDF object, using the response object as its "file."
#     buffer = BytesIO()
#     p = canvas.Canvas(buffer)

#     # Draw things on the PDF. Here's where the PDF generation happens.
#     # See the ReportLab documentation for the full list of functionality.
#     p.drawString(100, 100, "Hello world.")

#     # Close the PDF object cleanly, and we're done.
#     p.showPage()
#     p.save()

#     # FileResponse sets the Content-Disposition header so that browsers
#     # present the option to save the file.
#     buffer.seek(0)
#     return FileResponse(buffer, as_attachment=True, filename='hello.pdf')



from django.shortcuts import render
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from course.models import Course

def generate_course_pdf(request):
    courses = Course.objects.all()

    # create a Django response object, and specify content type as pdf
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="all_courses.pdf"'

    # find the template and render it
    template = get_template('all_courses.html')
    html = template.render({'courses': courses})

    # create a pdf
    pisa_status = pisa.CreatePDF(html, dest=response)

    # if error then show some funny view
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')

    return response

# def generate_pdf(template_src, context_dict, file_name='aceapp-pdf'):
#     # Create a Django response object, and specify content type as PDF
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename={file_name}'

#     # Find the template and render it.
#     template = get_template(template_src)
#     html = template.render(context_dict)

#     # Create a PDF
#     pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)

#     # If there was an error, return a response with an error message
#     if pisa_status.err:
#         return HttpResponse('We had some errors <pre>' + html + '</pre>')

#     return response


# from rest_framework import viewsets
# class GeneratePdfViewset(viewsets.ViewSet):
#     response =generate_pdf('temp/sample_template.html',
#                 {
#                     'name': "name_of _the pdf"
#                 }, 'name _of _the pdf.pdf')
#     return response



import pandas as pd

@api_view(['POST'])
def convert_to_excel(request):
    # Get the template from the request data
    template = request.data['template']

    # Convert the template to Excel using pandas
    table = pd.read_html(template)[0]
    excel_writer = pd.ExcelWriter('output.xlsx')
    table.to_excel(excel_writer, index=False)
    excel_writer.save()

    # Return the Excel data as a response
    with open('output.xlsx', 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="my_excel_file.xlsx"'
        return response








#add field based data 
# def get_table_data(model, field, filter_field=None, filter_value=None):
#     # Get all fields in the model
#     fields = model._meta.get_fields()

#     # Extract the names of fields that are not related fields or ManyToMany fields
#     headers = [field.name.replace('_', ' ').title() for field in fields
#                if isinstance(field, models.Field) and not (field.one_to_one or field.many_to_many)
#                and field.name != 'photo' and field.name != 'description' and field.name != 'modeofclasschoice'
#                and field.name != 'joined_date' and field.name != 'last_login' and field.name != 'is_staff'
#                and field.name != 'is_active' and field.name != 'is_verified' and field.name != 'is_blocked'
#                and field.name != 'blockreason' and field.name != 'is_rejected' and field.name != 'rejectreason'
#                and field.name != 'inhouse_fac' and field.name != 'identity_card' and field.name != 'resume'
#                and field.name != 'experiance_link' and field.name != 'identity_card' and field.name != 'password' and field.name != 'is_faculty' and field.name != 'is_roleuser' and field.name != 'is_superuser']

#     # Filter instances based on the filter field and value
#     filter_kwargs = {}
#     if filter_field and filter_value:
#         filter_kwargs = {filter_field: filter_value}
#     instances = model.objects.filter(**filter_kwargs)

#     # Extract the data for each instance
#     data = []
#     for instance in instances:
#         row = []
#         for header in headers:
#             value = getattr(instance, header.replace(' ', '_').lower(), None)
#             if value is not None:
#                 # Check if the value is a string and longer than 14 characters
#                 if isinstance(value, str) and len(value) > 14:
#                     # Split the string into chunks of 14 characters and add a newline character after each chunk
#                     value = '\n'.join([value[i:i+14] for i in range(0, len(value), 14)])
#                 row.append(value)
#             else:
#                 row.append('')
#         data.append(row)
#         print(data,'return data ')

#     return headers, data

########DDD
def get_table_data(model, field, filter_field=None, filter_value=None):
    # Get all fields in the model
    fields = model._meta.get_fields()

    # Extract the names of fields that are not related fields or ManyToMany fields
    headers = [field.name.replace('_', ' ').title() for field in fields
               if isinstance(field, models.Field) and not (field.one_to_one or field.many_to_many)
               and field.name != 'photo' and field.name != 'description' and field.name != 'modeofclasschoice'
               and field.name != 'joined_date' and field.name != 'last_login' and field.name != 'is_staff'
               and field.name != 'is_active' and field.name != 'is_verified' and field.name != 'is_blocked'
               and field.name != 'blockreason' and field.name != 'is_rejected' and field.name != 'rejectreason'
               and field.name != 'inhouse_fac' and field.name != 'identity_card' and field.name != 'resume'
               and field.name != 'experiance_link' and field.name != 'identity_card' and field.name != 'password' and field.name != 'is_faculty' and field.name != 'is_roleuser' and field.name != 'is_superuser'
               and field.name != 'photoverified' and field.name != 'resumeverified' and field.name != 'idverified' and field.name!='otp']

    # Filter instances based on the filter field and value
    filter_kwargs = {}
    if filter_field and filter_value:
        filter_kwargs = {filter_field: filter_value}
    instances = model.objects.filter(**filter_kwargs)

    # Extract the data for each instance
    data = []
    for instance in instances:
        row = []
        for header in headers:
            # Get the value of the attribute
            value = getattr(instance, header.replace(' ', '_').lower(), None)
            if value is not None:
                # Check if the value is an object and has a string representation
                if isinstance(value, models.Model) and hasattr(value, 'name'):
                    # Set the value to the object name
                    value = value.name
                # Check if the value is a string and longer than 14 characters
                elif isinstance(value, str) and len(value) > 14:
                    # Split the string into chunks of 14 characters and add a newline character after each chunk
                    value = '\n'.join([value[i:i+14] for i in range(0, len(value), 14)])
                row.append(value)
            else:
                row.append('')
        data.append(row)
        print(data,'return data ')

    return headers, data 





##jjjjjjj########
import textwrap
from course.views import getallNew
class GeneratePdfViewset(viewsets.ViewSet):
    def list(self, request):
        model_name = request.GET.get('model', None)
        app_label = request.GET.get('app_label', None)
        filter_field = request.GET.get('filter_field', None)
        filter_value = request.GET.get('filter_value', None)
        course_id=request.GET.get('course_id',None)

        if course_id:
            
            course = get_object_or_404(Course, pk=course_id)
            subjects = Subject.objects.filter(course=course)
            modules = Module.objects.filter(subject__in=subjects)
            topics = Topic.objects.filter(module__in=modules)
            subtopics = SubTopic.objects.filter(topic__in=topics)
            print(subtopics,'subtopic###')
            
            # Create a DataFrame with the relevant information
            data = []
            prev_course_name = ""
            prev_subject_name = ""
            prev_module_name = ""
            prev_topic_name = ""
            

            for subtopic in subtopics:
                course_name = subtopic.topic.module.subject.course.name
                course_name = '\n'.join(textwrap.wrap(course_name, width=14))

                subject_name = subtopic.topic.module.subject.name
                subject_name = '\n'.join(textwrap.wrap(subject_name, width=14))

                module_name = subtopic.topic.module.name
                module_name = '\n'.join(textwrap.wrap(module_name, width=14))

                topic_name = subtopic.topic.name
                topic_name = '\n'.join(textwrap.wrap(topic_name, width=14))


                
                if course_name == prev_course_name:
                    course_name = ""
                else:
                    prev_course_name = course_name
                
                if subject_name == prev_subject_name:
                    subject_name = ""
                else:
                    prev_subject_name = subject_name
                
                if module_name == prev_module_name:
                    module_name = ""
                else:
                    prev_module_name = module_name
                
                if topic_name == prev_topic_name:
                    topic_name = ""
                else:
                    prev_topic_name = topic_name
                
                # Convert time from minutes to hours and minutes
                time_in_minutes = subtopic.time_needed
                hours, minutes = divmod(time_in_minutes, 60)

                # Format time as "0hrs 10min"
                time_string = f"{hours} hrs {minutes} min"

                data.append({
                    'Course': course_name,
                    'Subject': subject_name,
                    'Module': module_name,
                    'Topic': topic_name,
                    'Subtopic': subtopic.name,
                    'SubtopicTime': time_string
})
        
            print(data,'18888888888888888888888888')
            nameheading='Course'
            current_datetime = timezone.now()
            resp = generate_pdf('coursepdf.html',
                                {
                                    # 'headers': headers,
                                    'data': data,
                                    'current_datetime': current_datetime,
                                    'model':nameheading
                                }, 'courselist.pdf')
            return resp

        elif app_label and model_name:
            model = apps.get_model(app_label=app_label, model_name=model_name)
            headers, data = get_table_data(model, field='*', filter_field=filter_field, filter_value=filter_value)

            current_datetime = timezone.now()
            resp = generate_pdf('commonpdf.html',
                                {
                                    'headers': headers,
                                    'data': data,
                                    'model': model_name,
                                    'current_datetime': current_datetime
                                }, 'ace-app.pdf')
            return resp

        else:
            return Response({"error": "Invalid request parameters"})

import openpyxl
from django.http import HttpResponse
from django.db.models import Q
from accounts.models import Faculty

def export_blocked_faculties_to_excel(request):
    # Get all the Faculty objects where is_blocked=True
    faculties = Faculty.objects.filter(is_blocked=True)

    # Create a new workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add headers to the sheet
    sheet.append(['Faculty Name', 'Address','PhoneNumber','Qualification'])

    # Add data to the sheet
    for faculty in faculties:
        sheet.append([faculty.name, faculty.address,faculty.user.mobile,faculty.qualification])

    # Create a response object with the Excel file
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="blocked_faculties.xlsx"'

    # Save the workbook to the response object
    workbook.save(response)

    return response

def export_applied_faculties_to_excel(request):
    # Get all the Faculty objects where is_blocked=True
    faculties = Faculty.objects.filter(is_verified=False)

    # Create a new workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add headers to the sheet
    sheet.append(['Faculty Name', 'Address','PhoneNumber','Qualification'])

    # Add data to the sheet
    for faculty in faculties:
        sheet.append([faculty.name, faculty.address,faculty.user.mobile,faculty.qualification])


    # Create a response object with the Excel file
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Applied_faculties.xlsx"'

    # Save the workbook to the response object
    workbook.save(response)

    return response

def export_active_faculties_to_excel(request):
    # Get all the Faculty objects where is_blocked=True
    faculties = Faculty.objects.filter(is_verified=True)

    # Create a new workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add headers to the sheet
    sheet.append(['Faculty Name', 'Address','PhoneNumber','Qualification'])

    # Add data to the sheet
    for faculty in faculties:
        sheet.append([faculty.name, faculty.address,faculty.user.mobile,faculty.qualification])


    # Create a response object with the Excel file
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Active_faculties.xlsx"'

    # Save the workbook to the response object
    workbook.save(response)

    return response
def export_rejected_faculties_to_excel(request):
    # Get all the Faculty objects where is_blocked=True
    faculties = Faculty.objects.filter(is_rejected=True)

    # Create a new workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add headers to the sheet
    sheet.append(['Faculty Name', 'Address','PhoneNumber','Qualification'])

    # Add data to the sheet
    for faculty in faculties:
        sheet.append([faculty.name, faculty.address,faculty.user.mobile,faculty.qualification])


    # Create a response object with the Excel file
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Rejected_faculties.xlsx"'

    # Save the workbook to the response object
    workbook.save(response)
    return response


from openpyxl import Workbook

def export_holiday_list(request):
    try:
         # create a new workbook
        wb = Workbook()

        # select the active worksheet
        ws = wb.active

        # define the header row
        ws.append(['Name', 'Date', 'Levels', 'Batches', 'Branches'])

        # get all the special holidays
        special_holidays = SpecialHoliday.objects.all()

        # loop through the special holidays
        for special_holiday in special_holidays:
            # get the levels, batches, and branches associated with the special holiday
            levels = special_holiday.levels.all()
            batches = special_holiday.batches.all()
            branches = special_holiday.branches.all()

            # create a list of level names
            level_names = [level.name for level in levels]

            # create a list of batch names
            batch_names = [batch.name for batch in batches]

            # create a list of branch names
            branch_names = [branch.name for branch in branches]

            # add a row to the worksheet for the special holiday
            ws.append([
                special_holiday.name,
                special_holiday.date,
                ', '.join(level_names),
                ', '.join(batch_names),
                ', '.join(branch_names),
            ])

        # create a file-like object for the response
        output = io.BytesIO()

        # save the workbook to the file-like object
        wb.save(output)

        # set the file pointer at the beginning of the file-like object
        output.seek(0)

        # set the response content type to Excel
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # set the filename for the response
        response['Content-Disposition'] = 'attachment; filename=SpecialHolidays.xlsx'

        # return the response
        return response

    except ValidationError as e:
        error_msg = 'Validation error: ' + str(e)
        return JsonResponse({'error': error_msg}, status=400)

    except Exception as e:
        error_msg = 'Internal Server Error: ' + str(e)
        return JsonResponse({'error': error_msg}, status=400)

# def generate_excel_for_course(request, course_id):
#     course = get_object_or_404(Course, pk=course_id)
    
#     # Query all related objects for the given course
#     subjects = Subject.objects.filter(course=course)
#     modules = Module.objects.filter(subject__in=subjects)
#     topics = Topic.objects.filter(module__in=modules)
#     subtopics = SubTopic.objects.filter(topic__in=topics)
    
#     # Create a DataFrame with the relevant information
#     data = []
#     for subtopic in subtopics:
#         data.append({
#             'Course': course.name,
#             'Subject': subtopic.topic.module.subject.name,
#             'Module': subtopic.topic.module.name,
#             'Topic': subtopic.topic.name,
#             'Subtopic': subtopic.name,
#             'TimeNeeded': subtopic.time_needed
#         })
#     df = pd.DataFrame(data)
    
#     # Write the DataFrame to an Excel file and return it as an HttpResponse
#     response = HttpResponse(content_type='application/vnd.ms-excel')
#     response['Content-Disposition'] = f'attachment; filename={course.name}.xlsx'
#     df.to_excel(response, index=False)
#     return response

# def generate_excel_for_course(request, course_id):
#     course = get_object_or_404(Course, pk=course_id)
    
#     # Query all related objects for the given course
#     subjects = Subject.objects.filter(course=course)
#     modules = Module.objects.filter(subject__in=subjects)
#     topics = Topic.objects.filter(module__in=modules)
#     subtopics = SubTopic.objects.filter(topic__in=topics)
    
#     # Create a DataFrame with the relevant information
#     data = []
#     prev_course_name = ""
#     prev_subject_name = ""
#     prev_module_name = ""
#     prev_topic_name = ""
    
#     for subtopic in subtopics:
#         course_name = subtopic.topic.module.subject.course.name
#         subject_name = subtopic.topic.module.subject.name
#         module_name = subtopic.topic.module.name
#         topic_name = subtopic.topic.name
        
#         if course_name == prev_course_name:
#             course_name = ""
        
#         if subject_name == prev_subject_name:
#             subject_name = ""
        
#         if module_name == prev_module_name:
#             module_name = ""
        
#         if topic_name == prev_topic_name:
#             topic_name = ""
        
#         data.append({
#             'Course': course_name,
#             'Subject': subject_name,
#             'Module': module_name,
#             'Topic': topic_name,
#             'Subtopic': subtopic.name,
#             'SubtopicTime': subtopic.time_needed
#         })
        
#         prev_course_name = subtopic.topic.module.subject.course.name
#         prev_subject_name = subtopic.topic.module.subject.name
#         prev_module_name = subtopic.topic.module.name
#         prev_topic_name = subtopic.topic.name
        
#     df = pd.DataFrame(data)
    
#     # Write the DataFrame to an Excel file and return it as an HttpResponse
#     response = HttpResponse(content_type='application/vnd.ms-excel')
#     response['Content-Disposition'] = f'attachment; filename={course.name}.xlsx'
#     df.to_excel(response, index=False)
#     return response

# def generate_excel_for_course(request, course_id):
#     course = get_object_or_404(Course, pk=course_id)
    
#     # Query all related objects for the given course
#     subjects = Subject.objects.filter(course=course)
#     modules = Module.objects.filter(subject__in=subjects)
#     topics = Topic.objects.filter(module__in=modules)
#     subtopics = SubTopic.objects.filter(topic__in=topics)
    
#     # Create a DataFrame with the relevant information
#     data = []
#     for subtopic in subtopics:
#         data.append({
#             'Course': course.name if subtopic.topic.module.subject.course == course else '',
#             'Subject': subtopic.topic.module.subject.name if subtopic.topic.module.subject in subjects else '',
#             'Module': subtopic.topic.module.name if subtopic.topic.module in modules else '',
#             'Topic': subtopic.topic.name,
#             'Subtopic': subtopic.name,
#             'SubtopicTime': subtopic.time_needed
#         })
#     df = pd.DataFrame(data)
    
#     # Write the DataFrame to an Excel file and return it as an HttpResponse
#     response = HttpResponse(content_type='application/vnd.ms-excel')
#     response['Content-Disposition'] = f'attachment; filename={course.name}.xlsx'
#     df.to_excel(response, index=False)
#     return response



# def generate_excel_for_course(request, course_id):
#     course = get_object_or_404(Course, pk=course_id)
    
#     # Query all related objects for the given course
#     subjects = Subject.objects.filter(course=course).order_by('priority')
#     modules = Module.objects.filter(subject__in=subjects).order_by('priority')
#     topics = Topic.objects.filter(module__in=modules).order_by('priority')
#     subtopics = SubTopic.objects.filter(topic__in=topics).order_by('priority')
    
#     # Create a DataFrame with the relevant information
#     data = []
#     prev_course_name = ""
#     prev_subject_name = ""
#     prev_module_name = ""
#     prev_topic_name = ""
    
#     for subtopic in subtopics:
#         course_name = subtopic.topic.module.subject.course.name
#         subject_name = subtopic.topic.module.subject.name
#         module_name = subtopic.topic.module.name
#         topic_name = subtopic.topic.name
        
#         if course_name == prev_course_name:
#             course_name = ""
#         else:
#             prev_course_name = course_name
        
#         if subject_name == prev_subject_name:
#             subject_name = ""
#         else:
#             prev_subject_name = subject_name
        
#         if module_name == prev_module_name:
#             module_name = ""
#         else:
#             prev_module_name = module_name
        
#         if topic_name == prev_topic_name:
#             topic_name = ""
#         else:
#             prev_topic_name = topic_name
        
#         data.append({
#             'Course': course_name,
#             'Subject': subject_name,
#             'Module': module_name,
#             'Topic': topic_name,
#             'Subtopic': subtopic.name,
#             'SubtopicTime': subtopic.time_needed
#         })
#         print(data,'daaaaaaaaaaaata')
#     df = pd.DataFrame(data)
    
#     # Write the DataFrame to an Excel file and return it as an HttpResponse
#     response = HttpResponse(content_type='application/vnd.ms-excel')
#     response['Content-Disposition'] = f'attachment; filename={course.name}.xlsx'
#     df.to_excel(response, index=False)
#     return response

def generate_excel_for_course(request, course_id):
    course = get_object_or_404(Course, pk=course_id)
    
    # Query all related objects for the given course
    # subjects = Subject.objects.filter(course=course).order_by('course','priority')
    # modules = Module.objects.filter(subject_in=subjects).order_by('subject_priority','priority')
    # topics = Topic.objects.filter(module__in=modules).order_by('priority')
    # subtopics = SubTopic.objects.filter(topic__in=topics).order_by('priority')
    
    # Create a DataFrame with the relevant information
    data = []
    prev_course_name = ""
    prev_subject_name = ""
    prev_module_name = ""
    prev_topic_name = ""
    new=[]
    subjects = Subject.objects.filter(course=course).order_by('priority')
    for subject in subjects:

        modules = Module.objects.filter(subject=subject).order_by('priority')
        for module in modules:

            topics = Topic.objects.filter(module=module).order_by('priority')
            for topic in topics:

                subtopics = SubTopic.objects.filter(topic=topic).order_by('priority')
                for subtopic in subtopics:
                    temp={"course_name" : subtopic.topic.module.subject.course.name,
                            "subject_name" : subtopic.topic.module.subject.name,
                            "module_name" :subtopic.topic.module.name,
                            "topic_name": subtopic.topic.name,
                            "name": subtopic.name,
                            "time_needed":subtopic.time_needed}
                    new.append(temp)




    
    for subtopic in new:
        course_name = subtopic["course_name"]
        subject_name = subtopic["subject_name"]
        module_name = subtopic["module_name"]
        topic_name = subtopic["topic_name"]
        
        if course_name == prev_course_name:
            course_name = ""
        else:
            prev_course_name = course_name
        
        if subject_name == prev_subject_name:
            subject_name = ""
        else:
            prev_subject_name = subject_name
        
        if module_name == prev_module_name:
            module_name = ""
        else:
            prev_module_name = module_name
        
        if topic_name == prev_topic_name:
            topic_name = ""
        else:
            prev_topic_name = topic_name
        
        data.append({
            'Course': course_name,
            'Subject': subject_name,
            'Module': module_name,
            'Topic': topic_name,
            'Subtopic': subtopic["name"],
            'SubtopicTime': subtopic["time_needed"]
        })
    df = pd.DataFrame(data)
    
    # Write the DataFrame to an Excel file and return it as an HttpResponse
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename={course.name}.xlsx'
    df.to_excel(response, index=False)
    return response


from reportlab.pdfgen import canvas

from django.http import HttpResponse
from io import BytesIO

def queryset_to_pdf(queryset, field_names):
    buffer = BytesIO()

    # Create a new PDF document
    pdf = canvas.Canvas(buffer)

    # Set up the coordinates for the content
    y = 800  # Initial y-coordinate
    line_height = 20

    # Iterate over the queryset and write the data to the PDF
    for obj in queryset:
        # Get the field values from the queryset object based on the field names
        field_values = [getattr(obj, field_name) for field_name in field_names]

        # Write the fields to the PDF at the specified coordinates
        x = 10  # Initial x-coordinate
        for field_value in field_values:
            pdf.drawString(x, y, str(field_value))
            x += 15  # Increase the x-coordinate for the next field

        # Increment the y-coordinate for the next row
        y -= line_height

    # Save and close the PDF document
    pdf.save()

    # Set the buffer's file pointer at the beginning
    buffer.seek(0)

    return buffer

import openpyxl
from django.http import HttpResponse

def queryset_to_excel(queryset, fields, k=None):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write headers
    headers = list(fields)
    sheet.append(headers)

    # Write data rows
    for obj in queryset:
        row = []
        for field in headers:
            
            try:
                print(k['is_online'][True],"kkkkkkkkkkkkkkkk")
                print(field,"fffffffffffffffffffff")
                if k and field in k and str(getattr(obj, field)) in k[field]:
                    row.append(str(k[field][str(getattr(obj, field))]))
                else:
                    if field=='batch':
                        row.append(str(getattr(obj, field).name))
                    elif field=='branch':
                        row.append(str(getattr(obj, field).name))
                    else:
                        row.append(str(getattr(obj, field)))


                    # row.append(str(obj.field))

            except Exception as e:
                print(type(e),"00000000000000000000000000")
                pass
        sheet.append(row)

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=your_data.xlsx'

    workbook.save(response)
    return response

def queryset_to_excel_data(queryset, fields, k={},manyKey={},arr=[]):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    model = queryset.model  # Get the model class from the queryset
    model_meta = model._meta
    
    keys={}
    for field in model_meta.get_fields():
        keys[field.name]=str(field.get_internal_type())
        if field.is_relation:
            relationship_type = field.get_internal_type()
            print(f"Field: {field.name}, Relationship Type: {relationship_type}")
    print(keys)

    def replacing_values(field,j):
        try:

            print(j,"ppppppppppppppppp",k[field],"iiiiiiiiiiiiiiiiii",field)

            if  j in k[field] :
                return str(k[field][j])
            else:
                return str(j)
        except:
            return str(j)
                
    # Write headers
    # headers = list(fields)
    # sheet.append(headers)
    if fields is None:
        headers = [field.name for field in model._meta.fields]
    else:
        headers = fields
    sheet.append(headers)


    data =[]
    # Write data rows
    for obj in queryset:
        obj_data = []
        
        for field in headers:
            # if k and field in k and str(getattr(obj, field)) in k[field]:
            #         obj_data.append(str(k[field][str(getattr(obj, field))]))
            # else:
            #     if field=='batch':
            #         obj_data.append(str(getattr(obj, field).name))
            #     elif field=='branch':
            #         obj_data.append(str(getattr(obj, field).name))
            #     else:
            #         obj_data.append(str(getattr(obj, field)))

            try:
                new=keys[field]
            except:
                new = None

            if new=='ManyToManyField':

                if str(field) in arr:
                     
                    j=''
                    
                    val=getattr(obj, field, None)
                    p=val.all()
                    for value in p:

                        j+=str(value.__dict__.get(manyKey[field], None)) + ', '

                    obj_data.append(replacing_values(field,j))

                else:
                     
                    j=''
                    val=getattr(obj, field, None)
                    p=val.all()
                    for value in p:
                        # j+=str(value.name) + ', '
                        j+=str(value.__dict__.get('name', None)) + ', '

                    obj_data.append(replacing_values(field,j))
                
                
            elif '__' in field:  # Check if the field has a "__" indicating a related field
                related_fields = field.split('__')
                value = obj
                for related_field in related_fields:
                    value = getattr(value, related_field, None)
                    if value is None:
                        break
                obj_data.append(replacing_values(field,value))
                
            
            else:
                value = getattr(obj, field, None)
                obj_data.append(replacing_values(field,value))
               
                
                    
        
        sheet.append(obj_data)
    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=your_data.xlsx'

    workbook.save(response)
    return response






from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors

def queryset_to_pdf(queryset, fields):
    # Create a response object with PDF content type
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=your_data.pdf'

    # Create a PDF document object using response as the file
    doc = SimpleDocTemplate(response, pagesize=letter)

    # Create a list for holding table data
    data = [fields]

    # Add data rows from queryset
    for obj in queryset:
        row = [str(getattr(obj, field)) for field in fields]
        data.append(row)

    # Create table object and set table style
    table = Table(data)
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                               ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                               ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                               ('FONTSIZE', (0, 0), (-1, 0), 14),
                               ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                               ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                               ('GRID', (0, 0), (-1, -1), 1, colors.black)]))

    # Build the PDF document
    elements = [table]
    doc.build(elements)

    return response

@api_view(['GET'])
def export_to_pdf(request):
    queryset = Faculty.objects.all()
    fields = ['id', 'name', 'address','whatsapp_contact_number','date_of_birth','qualification','district']  # Replace with your desired fields

    response = queryset_to_pdf(queryset, fields)

    return response

def export_inhouse_faculties_to_excel(request):
    # Get all the Faculty objects where is_blocked=True
    faculties = Faculty.objects.filter(inhouse_fac=True)

    # Create a new workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add headers to the sheet
    sheet.append(['Faculty Name', 'Address','PhoneNumber','Qualification'])

    # Add data to the sheet
    for faculty in faculties:
        sheet.append([faculty.name, faculty.address,faculty.user.mobile,faculty.qualification])
    # Create a response object with the Excel file
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Inhouse_faculties.xlsx"'

    # Save the workbook to the response object
    workbook.save(response)

    return response

def export_offline_faculties_to_excel(request):
    # Get all the Faculty objects where is_blocked=True
    faculties = Faculty.objects.filter(modeofclasschoice=1)

    # Create a new workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add headers to the sheet
    sheet.append(['Faculty Name', 'Address','PhoneNumber','Qualification'])

    # Add data to the sheet
    for faculty in faculties:
        sheet.append([faculty.name, faculty.address,faculty.user.mobile,faculty.qualification])
    # Create a response object with the Excel file
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Offline_faculties.xlsx"'

    # Save the workbook to the response object
    workbook.save(response)

    return response


def export_online_faculties_to_excel(request):
    # Get all the Faculty objects where is_blocked=True
    faculties = Faculty.objects.filter(modeofclasschoice=2)

    # Create a new workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add headers to the sheet
    sheet.append(['Faculty Name', 'Address','PhoneNumber','Qualification'])

    # Add data to the sheet
    for faculty in faculties:
        sheet.append([faculty.name, faculty.address,faculty.user.mobile,faculty.qualification])
    # Create a response object with the Excel file
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Online_faculties.xlsx"'

    # Save the workbook to the response object
    workbook.save(response)

    return response

